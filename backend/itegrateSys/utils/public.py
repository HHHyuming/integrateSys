from io import BytesIO
from openpyxl import Workbook, load_workbook


class ExcelInterface(object):

    def __init__(self, model=None, sheet_name='test'):
        self._model = model
        self.sheet_name = sheet_name

    def write_from_model(self):
        wb = Workbook()
        ws = wb.active
        wb.create_sheet(self.sheet_name)
        # 写入表头
        ws.append(['用户名', '性别', '邮箱', '部门', '状态', '创建时间'])
        ws.append(['yuming', 'male', '2219659698@qq.com', 'development', 'active', '2020-20-20-20'])
        wb.save('./test.xlsx')

    def get_xlsx_data(self):
        with open('./test.xlsx', 'rb') as f:
            result = f.read()
        return BytesIO(result)
